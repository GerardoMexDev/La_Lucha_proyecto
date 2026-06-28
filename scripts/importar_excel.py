#!/usr/bin/env python3
"""
Importación única del Excel de control contable a la base de datos de la app.

Uso:
    python scripts/importar_excel.py            # importa
    python scripts/importar_excel.py --dry-run  # previsualiza sin insertar
"""

import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "Control_Contable_Sueldos_Gomeria_Monedas_2.xlsx"
DB_PATH    = Path(__file__).parent.parent / "instance" / "caja_chica.db"

BLOCK_SIZE = 11   # 9 columnas de datos + 2 separadoras vacías
NUM_BLOCKS = 6    # semanas del Excel

# --------------------------------------------------------------------------- #
# Normalización de método de pago
# --------------------------------------------------------------------------- #
METODO_MAP = {
    "tc":            "Tarjeta",
    "tarjeta":       "Tarjeta",
    "tarjeta brou":  "Tarjeta",
    "post":          "Tarjeta",
    "transferencia": "Transferencia",
    "tranferencia":  "Transferencia",
    "cheque":        "Cheque",
    "efectivo":      "Efectivo",
}

def normalizar_metodo(raw):
    if not raw:
        return "Efectivo"
    clave = str(raw).strip().lower()
    return METODO_MAP.get(clave, "Efectivo")


# --------------------------------------------------------------------------- #
# Filas que se omiten siempre (aperturas de caja = saldo de arrastre)
# --------------------------------------------------------------------------- #
RE_OMITIR = re.compile(
    r"^(inicio\s*(de\s*)?caja?|inicio\s*campa[ñn]a)$",
    re.IGNORECASE
)


# --------------------------------------------------------------------------- #
# Mapeo concepto → categoría, separado por tipo
# --------------------------------------------------------------------------- #
PATRONES_INGRESO = [
    (re.compile(r"venta|cubierta|c[aá]mara|tarina|tarinas|compra\s+cubierta|compra\s+2\s+cubiert"
                r"|pago\s+cubierta|cuota\s+cubierta|saldo\s+cubierta", re.I), "Ventas"),
    (re.compile(r"trabajo|pinchazo|auxilio|balanceo|reparaci[oó]n\s+cubierta|servicio"
                r"|desenllante|llanta|parche|talon|trabao|valvula|camb.*llanta", re.I),
     "Servicios de taller"),
    # cobros y pagos de clientes
    (re.compile(r"cobr|pago\s+agrotech|pago\s+leo|pago\s+marco|pago\s+daniel|pago\s+manga"
                r"|pago\s+vicera|pago\s+cubiertas|cobranza|pago\s+trabajos|pago\s+de\b", re.I),
     "Otros ingresos"),
]

PATRONES_EGRESO = [
    (re.compile(r"adelant[aeo]|anticipo|adellanto|adelante", re.I), "Adelanto de sueldo"),
    (re.compile(r"sueldo|finiquito", re.I),                         "Sueldos y adelantos"),
    (re.compile(r"vi[aá]tico|viaje\s+a|gasolna|nafta\s+mandad|gasolina\s+mandad"
                r"|mandado|viaticos\s+salida", re.I),               "Viáticos"),
    (re.compile(r"nafta$|gasolina$|combustible|combutible|combustble", re.I), "Combustible"),
    (re.compile(r"retiro\s+renta|renta\s+local|alquiler", re.I),    "Alquiler local"),
    (re.compile(r"env[ií]o.*cheque|env[ií]os?\s+cheque|flete\s+cubiert|pago\s+env[ií]os", re.I),
     "Envíos"),
    (re.compile(r"retiro\s+dep[oó]sito|retiro\s*banco|retiro\s+d[oó]lar|retiro\s+de\s+d[oó]lar"
                r"|retiro\s+pesos|retiro\s+para|retiro$", re.I),    "Retiros"),
    (re.compile(r"papeler[ií]a|chip.*celular|recarga.*cel|recarga.*tel[eé]f|recibo\s+de\s+pago"
                r"|compra.*recibo", re.I),                           "Papelería / administrativo"),
    (re.compile(r"insumo|suministro|material|compra\s+material|repuesto|herramienta"
                r"|parche|cemento|refacci[oó]n|compra|sosa|chinches|tornillo|electrod"
                r"|cinta\s+aislante|detergente|aceite|disco\s|jabon|cascola|valvulina"
                r"|gasto\s+insumo|gasto\s+empresa", re.I),           "Insumos y suministros"),
    # carpintero, sorteo, reparaciones de infraestructura, etc.
    (re.compile(r"carpintero|sorteo|reparaci[oó]n\s+puerta|viaje|recarga\s+telefon", re.I),
     "Otros gastos"),
]


def to_float(val):
    """Convierte un valor de celda a float; devuelve 0 si no es numérico."""
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def inferir_categoria(concepto, tipo, cat_ids):
    patrones = PATRONES_INGRESO if tipo == "ingreso" else PATRONES_EGRESO
    fallback  = "Otros ingresos" if tipo == "ingreso" else "Otros gastos"

    for patron, nombre in patrones:
        if patron.search(concepto):
            if nombre in cat_ids:
                return cat_ids[nombre], nombre

    return cat_ids.get(fallback), fallback


# --------------------------------------------------------------------------- #
# Importación principal
# --------------------------------------------------------------------------- #
def importar(dry_run=False):
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encontró el Excel:\n  {EXCEL_PATH}")
        return
    if not DB_PATH.exists():
        print(f"ERROR: No se encontró la base de datos:\n  {DB_PATH}")
        print("Corré run.py al menos una vez para crearla.")
        return

    conn = sqlite3.connect(str(DB_PATH))
    cur  = conn.cursor()

    # Verificar si ya hay datos
    (total_existente,) = cur.execute("SELECT COUNT(*) FROM movimientos").fetchone()
    if total_existente > 0:
        print(f"AVISO: ya hay {total_existente} movimientos en la base de datos.")
        resp = input("¿Continuar igualmente? Esto sumará los datos del Excel a los existentes. [s/N] ").strip().lower()
        if resp != "s":
            print("Importación cancelada.")
            conn.close()
            return

    # Cargar ids de categorías
    cat_ids = {nombre: cid for cid, nombre in cur.execute("SELECT id, nombre FROM categorias")}

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["Caja Diaria"]

    insertados   = 0
    omitidos     = 0
    sin_monto    = 0
    por_categoria = {}

    for blk in range(NUM_BLOCKS):
        c0 = blk * BLOCK_SIZE
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            fecha_raw = row[c0 + 0]
            concepto  = row[c0 + 1]
            ing_usd   = to_float(row[c0 + 3])
            ing_uyu   = to_float(row[c0 + 4])
            egr_usd   = to_float(row[c0 + 5])
            egr_uyu   = to_float(row[c0 + 6])
            metodo    = row[c0 + 7]
            obs       = row[c0 + 8]

            if not fecha_raw or not concepto:
                continue
            concepto = str(concepto).strip()
            if not concepto or concepto.lower() == "concepto":
                continue
            if RE_OMITIR.match(concepto):
                omitidos += 1
                continue

            fecha = (fecha_raw.strftime("%Y-%m-%d")
                     if hasattr(fecha_raw, "strftime")
                     else str(fecha_raw)[:10])
            metodo_norm = normalizar_metodo(metodo)

            movs = []
            if ing_usd > 0:
                movs.append(("ingreso", ing_usd, "USD"))
            if ing_uyu > 0:
                movs.append(("ingreso", ing_uyu, "UYU"))
            if egr_usd > 0:
                movs.append(("egreso", egr_usd, "USD"))
            if egr_uyu > 0:
                movs.append(("egreso", egr_uyu, "UYU"))

            # Pagos en reales anotados en observaciones (R$NNN)
            if not movs and obs:
                m = re.search(r'R\$\s*(\d+(?:[.,]\d+)?)', str(obs))
                if m:
                    monto_brl = float(m.group(1).replace(',', '.'))
                    # El tipo se infiere del concepto: si es servicio/venta → ingreso
                    tipo_brl = "ingreso" if any(
                        p.search(concepto) for p, _ in PATRONES_INGRESO
                    ) else "egreso"
                    movs.append((tipo_brl, monto_brl, "BRL"))

            if not movs:
                sin_monto += 1
                continue

            for tipo, monto, moneda in movs:
                cat_id, cat_nombre = inferir_categoria(concepto, tipo, cat_ids)
                por_categoria[cat_nombre] = por_categoria.get(cat_nombre, 0) + 1

                if not dry_run:
                    cur.execute(
                        """INSERT INTO movimientos
                           (fecha, tipo, moneda, monto, concepto, categoria_id,
                            metodo_pago, observaciones)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (fecha, tipo, moneda, monto, concepto,
                         cat_id, metodo_norm,
                         str(obs) if obs else None)
                    )
                insertados += 1

    if not dry_run:
        conn.commit()
    conn.close()

    modo = "[DRY-RUN] " if dry_run else ""
    print(f"\n{modo}Resultado de la importación:")
    print(f"  Movimientos {'a insertar' if dry_run else 'insertados'}: {insertados}")
    print(f"  Filas omitidas (apertura de caja) : {omitidos}")
    print(f"  Filas sin monto válido             : {sin_monto}")
    print(f"\n  Por categoría:")
    for cat, n in sorted(por_categoria.items(), key=lambda x: -x[1]):
        print(f"    {cat:<35} {n:>4}")

    if dry_run:
        print("\n(no se insertó nada — corré sin --dry-run para importar)")
    else:
        print("\nListo. Abrí la app para revisar los datos.")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    importar(dry_run=dry)
